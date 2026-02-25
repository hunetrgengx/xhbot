#!/usr/bin/env python3
"""统计 bio_calls.jsonl 中 full_name 和 deleted_content 的出现次数，输出到 Excel"""

import json
from collections import Counter
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    exit(1)


def main():
    data_path = Path(__file__).parent / "bio_calls.jsonl"
    out_path = Path(__file__).parent / "bio_calls_stats.xlsx"

    full_name_counts = Counter()
    deleted_content_counts = Counter()
    bio_counts = Counter()
    full_name_len_counts = Counter()  # 用户名长度 -> 出现次数

    with open(data_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                data = json.loads(line)
                fn = data.get("full_name")
                if fn is not None:
                    full_name_counts[fn] += 1
                    full_name_len_counts[len(fn)] += 1

                dc = data.get("deleted_content")
                if dc is not None:
                    deleted_content_counts[dc] += 1
                elif data.get("bio") is not None:
                    bio_counts[data["bio"]] += 1
            except json.JSONDecodeError:
                continue

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: full_name 统计（按次数降序）
    ws_fn = wb.create_sheet("full_name 统计", 0)
    ws_fn.append(["full_name", "出现次数"])
    for name, cnt in full_name_counts.most_common():
        ws_fn.append([name, cnt])

    # Sheet 2: 用户名长度统计（按长度升序）
    ws_len = wb.create_sheet("用户名长度统计", 1)
    ws_len.append(["用户名长度", "出现次数"])
    for length in sorted(full_name_len_counts.keys()):
        ws_len.append([length, full_name_len_counts[length]])

    # Sheet 3: deleted_content 统计（按次数降序）
    ws_dc = wb.create_sheet("deleted_content 统计", 1)
    ws_dc.append(["deleted_content", "出现次数"])
    for content, cnt in deleted_content_counts.most_common():
        ws_dc.append([content, cnt])

    # Sheet 4: bio 统计（旧格式，有 bio 无 deleted_content 的记录）
    if bio_counts:
        ws_bio = wb.create_sheet("bio 统计", 3)
        ws_bio.append(["bio", "出现次数"])
        for content, cnt in bio_counts.most_common():
            ws_bio.append([content, cnt])

    # 自动列宽
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 80)

    wb.save(out_path)
    print(f"统计完成，已保存到: {out_path}")
    print(f"  full_name 去重数: {len(full_name_counts)}, 总记录: {sum(full_name_counts.values())}")
    if full_name_len_counts:
        print(f"  用户名长度分布: {len(full_name_len_counts)} 种长度, 最短 {min(full_name_len_counts)} 字, 最长 {max(full_name_len_counts)} 字")
    print(f"  deleted_content 去重数: {len(deleted_content_counts)}, 总记录: {sum(deleted_content_counts.values())}")
    if bio_counts:
        print(f"  bio 去重数: {len(bio_counts)}, 总记录: {sum(bio_counts.values())}")


if __name__ == "__main__":
    main()
